# from tempfile import NamedTemporaryFile
import openpyxl
from flask import current_app as app
from db.models import BrokerSource, LocalSample
from services import geo_localization_service
from .organism_service import get_or_create_organism
OPTIONS = ['SKIP','UPDATE']

def parse_excel(excel=None, id=None, taxid=None, scientific_name=None, header=1, option="SKIP", source=None):

##PARAMS VALIDATION
    param_errors=list()
    if not excel:
        e=dict()
        e['excel'] = ['excel field missing']
        param_errors.append(e)

    if not source:
        source = BrokerSource.LOCAL

    elif source not in [source.value for source in BrokerSource]:
        e=dict()
        e['source'] = [f'{source} is not present in sources']
        param_errors.append(e)

    if not isinstance(header,int):
        if not header.isnumeric():
            e=dict()
            e['header'] = [f"header must be a number. header: {header}"]
            param_errors.append(e)
        else:
            header = int(header)

    if header < 1:
        e = dict()
        e['header'] = ["header must be greater than 1"]
        param_errors.append(e)

    #can't continue if file is None
    if param_errors:
        return param_errors,400

    wb_obj = openpyxl.load_workbook(excel,data_only=True)
    sheet_obj = wb_obj.active

    #check if headers are correct
    header_row = [cell.value for cell in sheet_obj[header] if cell.value]
    if not header_row:
        e = dict()
        e['header'] = [f"header can't be greater than the total rows of the excel. header: {header}"]
        param_errors.append(e)

    mandatory_fields = [dict(key='Local Id', value=id), dict(key="taxid",value=taxid), dict(key="Scientific name",value=scientific_name)]

    for field in mandatory_fields:
        if not field['value']:
            e = dict()
            e[field['key']] = [f"{field['key']} field missing"]
            param_errors.append(e)
        elif not field['value'] in header_row:
            e = dict()
            e[field['key']] = [f"{field['value']} not found in {','.join(header_row)}"]
            param_errors.append(e)

    #check if options are valid
    if not option in OPTIONS:
        e = dict()
        e['import options'] = [f"option must be SKIP or UPDATE, current is: {option}"]
        param_errors.append(e)

    if param_errors:
        return param_errors,400


##EXCEL PARSING
    all_errors = list()
    saved_samples = list()
    for index, row in enumerate(list(sheet_obj.rows)[header:]):
        
        ##check if the row contains something, at least three cells
        values = len([cell.value for cell in row if cell.value])
        if values < 3:
            continue

        sample_error_obj=dict()
        sample_error_obj[index+header+1] = []
        new_sample = dict(metadata=dict())
        for key, cell in zip(header_row,row):
            if key in [f['value'] for f in mandatory_fields]:
                if not cell.value:
                    msg = f"{key} field is mandatory"
                    sample_error_obj[index+header+1].append(msg)
                else:
                    new_sample[key] = str(cell.value).strip()
            if cell.value:
                new_sample['metadata'][key] = cell.value
        
        if sample_error_obj[index+header+1]:
            all_errors.append(sample_error_obj)
            continue

        ##continue if errors are present
        if all_errors:
            continue

        saved_sample=dict()
        sample_obj = LocalSample.objects(local_id = new_sample[id]).first()
        if sample_obj:
            if option == 'SKIP':
                continue
            elif option == 'UPDATE':
                sample_obj.update(taxid=new_sample[taxid],local_id=new_sample[id],broker=source,metadata=new_sample['metadata'],scientific_name=new_sample[scientific_name])
                saved_sample[index+1+header] = [f"{sample_obj.local_id} correctly updated"]
        else:
            organism = get_or_create_organism(new_sample[taxid])
            if not organism:
                sample_error_obj[index+header+1] = 'TAXID not found in NCBI'
                all_errors.append(sample_error_obj)
                continue
            sample_obj = LocalSample(taxid=new_sample[taxid],local_id=new_sample[id],broker=source,metadata=new_sample['metadata'],scientific_name=new_sample[scientific_name]).save()                
            geo_localization_service.create_coordinates(sample_obj,organism)
            organism.local_samples.append(sample_obj.local_id)
            organism.save()
            saved_sample[index+1+header] = [f"{sample_obj.local_id} correctly saved"]
        saved_samples.append(saved_sample)

    if all_errors:
        return all_errors,400
    return saved_samples,201

    ##parse excel, return

# def get_sample_values(header,row):
#     values = {}
#     for key, cell in zip(header, row):
#         if cell.value:
#             values[key] = cell.value            
#     return values

# # return samples(rows)
# def parse_excel(excel, header_index):
#     wb_obj = openpyxl.load_workbook(excel,data_only=True)
#     sheet_obj = wb_obj.active
#     header = [cell.value.lower() for cell in sheet_obj[header_index] if cell.value]
#     #validations
#     return sheet_obj, header, header_index

# def create_excel():
#     samples = SecondaryOrganism.objects(accession=None).as_pymongo()
#     wb = Workbook()
#     ws = wb.active
#     ws.title='Metadata Entry'
#     ws.append(MANIFEST_HEADER)

#     for sample in samples:
#         parsed_sample = sample_to_excel(sample,MANIFEST_HEADER)
#         #convert fields 
#         ordered_list = list()
#         for field in MANIFEST_HEADER:
#             if field in parsed_sample.keys():
#                 ordered_list.append(parsed_sample[field])
#             else:
#                 ordered_list.append('')
#         ws.append(ordered_list)
#     with NamedTemporaryFile() as tmp:
#         wb.save(tmp.name)
#         return tmp.read()

# def sample_to_excel(sample, header):
#     formatted_sample=dict(TAXON_ID=sample['taxid'])
#     fields_converter = dict(
#         scientific_name='scientific_name',
#         geographic_location_country='collection_location',
#         geographic_location_region_and_locality='collection_location', 
#         collection_date='date_of_collection',
#         geographic_location_longitude='decimal_longitude',
#         geographic_location_latitude='decimal_latitude',
#         collecting_institution='collector_affiliation')
#     #format specific fields
#     for key,value in fields_converter.items():
#         if value == 'collection_location':
#             formatted_sample[value.upper()] = sample['geographic_location_country'] + ' | ' + sample['geographic_location_region_and_locality']
#         elif key in sample.keys():
#                 formatted_sample[value.upper()] = sample[key]
#     #format fields
#     for field in header:
#         if field.lower() in sample.keys():
#             formatted_sample[field] = sample[field.lower()]
#     return formatted_sample